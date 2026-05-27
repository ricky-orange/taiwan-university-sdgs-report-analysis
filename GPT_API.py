# -*- coding: utf-8 -*-
"""
SDG 永續報告書視覺分析工具 (Production Mode - v11)
- 依賴套件: pip install openai pdf2image openpyxl tenacity Pillow
- 依據 Standardized_Prompt_API.txt 的 10 欄格式輸出
"""

import base64
import csv
import io
import logging
import re
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import List, Optional

import openai
from openai import OpenAI
from pdf2image import convert_from_path
from openpyxl import Workbook, load_workbook
from tenacity import retry, wait_exponential, stop_after_attempt, retry_if_exception_type
from PIL import Image


# ==========================================
# 日誌設定
# ==========================================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s [%(levelname)s] %(message)s',
    datefmt='%H:%M:%S'
)
logger = logging.getLogger(__name__)


# ==========================================
# 1. 配置與常數
# ==========================================
@dataclass
class VisionConfig:
    # --- 檔案路徑 ---
    input_pdf: Path = Path("2024_NSYSU.pdf")
    output_xlsx: Path = Path("SDGs_analysis_NSYSU.xlsx")
    error_log: Path = Path("error_rows.log")
    api_key_path: Path = Path("API_Key.txt")
    prompt_path: Path = Path("Standardized_Prompt_API.txt")

    # --- 頁面範圍 ---
    start_page: int = 1
    end_page: int = 134

    # --- 模型與 API 參數 ---
    model_name: str = "gpt-5-mini"
    max_completion_tokens: int = 6000

    # --- 效能與影像控制 ---
    save_interval: int = 10          # 每 10 個有寫入資料的頁面存一次
    dpi_setting: int = 200
    image_max_size: int = 2048
    image_quality: int = 85

    # --- 固定輸出欄位 ---
    expected_columns: int = 10
    fixed_school_name: str = ""
    fixed_placeholder: str = "--"

    def __post_init__(self):
        if self.start_page < 1:
            raise ValueError("start_page 必須大於或等於 1")
        if self.end_page < self.start_page:
            raise ValueError("end_page 不能小於 start_page")
        # 從檔名取得學校縮寫
        parts = self.input_pdf.stem.split("_", 1)
        self.fixed_school_name = parts[1].upper() if len(parts) == 2 and parts[1] else ""


# ==========================================
# 2. 核心分析類別
# ==========================================
class SDGVisionAnalyzer:
    def __init__(self, config: VisionConfig):
        self.config = config
        self.client = self._setup_api()
        self.prompt_text = self._load_prompt()

    def _setup_api(self) -> OpenAI:
        if not self.config.api_key_path.exists():
            logger.error(f"找不到 API Key 檔案: {self.config.api_key_path}")
            sys.exit(1)

        with open(self.config.api_key_path, "r", encoding="utf-8") as f:
            api_key = f.read().strip()

        if not api_key:
            logger.error("API Key 檔案為空")
            sys.exit(1)

        return OpenAI(api_key=api_key)

    def _load_prompt(self) -> str:
        if not self.config.prompt_path.exists():
            logger.error(f"找不到 Prompt 檔案: {self.config.prompt_path}")
            sys.exit(1)

        with open(self.config.prompt_path, "r", encoding="utf-8") as f:
            prompt = f.read().strip()

        if not prompt:
            logger.error("Prompt 檔案內容為空")
            sys.exit(1)

        return prompt

    def _log_error(self, message: str):
        logger.warning(message.strip())
        with open(self.config.error_log, "a", encoding="utf-8") as f:
            f.write(message if message.endswith("\n") else message + "\n")

    def _process_image(self, image: Image.Image) -> str:
        img = image.copy()
        if max(img.size) > self.config.image_max_size:
            img.thumbnail(
                (self.config.image_max_size, self.config.image_max_size),
                Image.Resampling.LANCZOS
            )

        buffered = io.BytesIO()
        img.save(buffered, format="JPEG", quality=self.config.image_quality)
        return base64.b64encode(buffered.getvalue()).decode("utf-8")

    @retry(
        wait=wait_exponential(multiplier=2, min=4, max=60),
        stop=stop_after_attempt(5),
        retry=retry_if_exception_type((
            openai.RateLimitError,
            openai.APIConnectionError,
            openai.InternalServerError,
        )),
        reraise=True
    )
    def _call_ai_analysis(self, base64_image: str, page_num: int) -> str:
        final_user_content = self.prompt_text.replace(
            "{content}",
            "(請分析附圖中的文字內容)"
        )

        response = self.client.chat.completions.create(
            model=self.config.model_name,
            messages=[
                {
                    "role": "system",
                    "content": (
                        "你是一個專業的永續報告書內容分析師。"
                        "請嚴格遵守使用者提供的格式規範，只輸出純 CSV 或 NO_DATA。"
                        "不得輸出 Markdown、開場白、結尾說明。"
                    )
                },
                {
                    "role": "user",
                    "content": [
                        {"type": "text", "text": final_user_content},
                        {
                            "type": "image_url",
                            "image_url": {
                                "url": f"data:image/jpeg;base64,{base64_image}",
                                "detail": "high"
                            }
                        }
                    ]
                }
            ],
            max_completion_tokens=self.config.max_completion_tokens,
        )

        choice = response.choices[0]
        msg = choice.message
        content = msg.content

        if content:
            return content.strip()

        refusal = getattr(msg, "refusal", None)
        tool_calls = getattr(msg, "tool_calls", None)
        finish_reason = getattr(choice, "finish_reason", None)

        raise ValueError(
            f"模型未回傳文字內容 | finish_reason={finish_reason} | "
            f"refusal={refusal} | tool_calls={tool_calls}"
        )

    def _normalize_csv_text(self, text: str) -> str:
        marker = chr(96) * 3
        pattern = rf"{marker}(?:csv)?\n(.*?)\n{marker}"
        csv_match = re.search(pattern, text, re.DOTALL | re.IGNORECASE)
        raw_csv = csv_match.group(1).strip() if csv_match else text.strip()
        raw_csv.replace("\ufeff", "").strip()
    
        # 移除數字中的千分位逗號，例如 2,900 -> 2900、9,300 -> 9300
        raw_csv = re.sub(r'(?<=\d),(?=\d{3}(?!\d))', '', raw_csv)
        return raw_csv
    
    def _validate_and_clean_row(self, row: List[str], page_num: int) -> Optional[List[str]]:
        expected_cols = self.config.expected_columns

        # 1. 基本清洗
        row = [str(col).strip() for col in row]

        # 2. 略過空列
        if not row or not any(row):
            return None

        # 3. 欄位太少直接排除
        if len(row) < 5:
            return None

        # 4. 處理 9 欄情況：通常是少了 deviation_reason
        if len(row) == 9:
            last_val = row[-1].strip().lower()
            # 情況 A：disclosure_level 與 disclosure_reason 被合併
            m = re.match(r"^(高|中|低)[，,](.+)$", row[4].strip())

            if m:
                row[4] = m.group(1)
                row.insert(5, m.group(2).strip())

            # 情況 B：少了 deviation_reason
            elif "p." in last_val or last_val.isdigit():
                row.insert(8, self.config.fixed_placeholder)

        # 5. 嘗試拆解被合併的 --;-- 或 --；--
        if len(row) < expected_cols:
            for i, col in enumerate(row):
                if col in {"--;--", "--；--"}:
                    row = row[:i] + [self.config.fixed_placeholder, self.config.fixed_placeholder] + row[i + 1:]
                    if len(row) == expected_cols:
                        break

        # 6. 長度檢查
        if len(row) != expected_cols:
            return None

        # 7. 去除部分贅詞（沿用原始程式邏輯）
        redundant_patterns = r"^(本段(落)?(內容)?|此段(落)?|文中|段落(中)?|該圖|圖中)(提到|指出|顯示|描述|說明|表示|提及|主要|則|係)?(關於|指)?\s*[：,，]?"
        for idx in [2, 5]:  # sdgs_content, disclosure_reason
            clean_text = re.sub(redundant_patterns, "", row[idx]).strip()
            if clean_text and clean_text[0] in "，,：:":
                clean_text = clean_text[1:].strip()
            row[idx] = clean_text

        # 8. 去除校名主詞（沿用原始程式邏輯）
        school_name_patterns = r"^((國立|私立)?[\u4e00-\u9fff]{1,10}(科技大學|大學|學院)|本校|該校)"
        content = re.sub(school_name_patterns, "", row[2]).strip()
        if content and content[0] in "，,、；;：:":
            content = content[1:].strip()
        row[2] = content

        # 9. 固定欄位強制覆寫，完全對齊提示詞
        row[0] = self.config.fixed_school_name
        row[6] = self.config.fixed_placeholder
        row[7] = self.config.fixed_placeholder
        row[8] = self.config.fixed_placeholder
        row[9] = f"p.{page_num}"

        # 10. 必要欄位基本驗證
        sdgs_list = row[1].strip()
        if not sdgs_list:
            self._log_error(f"p.{page_num} 模型輸出空 sdgs_list 列，依規則此情況應回傳 NO_DATA: {row}")
            return None

        # 支援：
        # SDG4
        # SDG 4
        # SDG4；SDG11；SDG12
        # SDGs 1, 3, 5
        sdg_pattern = r"SDG[S]?\s*\d+(?:\s*[；;、,，]\s*(?:SDG[S]?\s*)?\d+)*"
        if not re.fullmatch(sdg_pattern, sdgs_list, re.IGNORECASE):
            return None

        # 11. disclosure_level 限定高/中/低
        if row[4] not in {"高", "中", "低"}:
            return None

        return row

    def _parse_and_validate(self, text: str, current_page: int) -> List[List[str]]:
        valid_rows: List[List[str]] = []
        raw_csv = self._normalize_csv_text(text)
        reader = csv.reader(io.StringIO(raw_csv))

        for row in reader:
            if not row or not any(row):
                continue

            first_cell = str(row[0]).strip().lower()
            if first_cell == "school_name":
                continue

            cleaned_row = self._validate_and_clean_row(row, current_page)
            if cleaned_row is None:
                self._log_error(f"p.{current_page} 格式異常或欄位不符: {row}")
                continue

            valid_rows.append(cleaned_row)

        return valid_rows

    def run(self):
        logger.info(
            f"開始任務：{self.config.input_pdf} "
            f"(p.{self.config.start_page} - p.{self.config.end_page})"
        )

        if not self.config.input_pdf.exists():
            logger.error(f"找不到 PDF 檔案: {self.config.input_pdf}")
            sys.exit(1)

        # 清空舊錯誤紀錄
        if self.config.error_log.exists():
            self.config.error_log.unlink()

        # 初始化 Excel
        if self.config.output_xlsx.exists():
            wb = load_workbook(self.config.output_xlsx)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.append([
                "school_name",
                "sdgs_list",
                "sdgs_content",
                "keywords",
                "disclosure_level",
                "disclosure_reason",
                "self_reported",
                "has_deviation",
                "deviation_reason",
                "page",
            ])

        total_written = 0
        unsaved_count = 0

        for current_page in range(self.config.start_page, self.config.end_page + 1):
            logger.info(f"分析 p.{current_page} ...")

            try:
                images = convert_from_path(
                    str(self.config.input_pdf),
                    first_page=current_page,
                    last_page=current_page,
                    dpi=self.config.dpi_setting
                )
                if not images:
                    self._log_error(f"p.{current_page} PDF 轉圖失敗: 未取得影像")
                    continue
                image = images[0]
            except Exception as e:
                self._log_error(f"p.{current_page} PDF 轉圖失敗: {e}")
                continue

            try:
                base64_image = self._process_image(image)
            finally:
                try:
                    image.close()
                except Exception:
                    pass
                del images

            try:
                result_text = self._call_ai_analysis(base64_image, current_page)
            except Exception as e:
                self._log_error(f"p.{current_page} API 徹底失敗: {e}")
                continue

            if not result_text or result_text.strip() == "NO_DATA":
                logger.info(f"  └ p.{current_page} 判定無資料")
                continue

            valid_rows = self._parse_and_validate(result_text, current_page)

            if valid_rows:
                for row in valid_rows:
                    ws.append(row)

                total_written += len(valid_rows)
                unsaved_count += 1
                logger.info(f"  └ 成功寫入 {len(valid_rows)} 筆")
            else:
                logger.info(f"  └ p.{current_page} 未取得有效格式資料")

            if unsaved_count >= self.config.save_interval:
                wb.save(self.config.output_xlsx)
                logger.info(f"--- 觸發批次存檔 (累積寫入 {total_written} 筆) ---")
                unsaved_count = 0

        wb.save(self.config.output_xlsx)

        logger.info(f"分析任務完成！總計寫入 {total_written} 筆有效資料。")
        if self.config.error_log.exists():
            logger.warning(f"部分資料格式異常被排除，請檢視 {self.config.error_log}")


if __name__ == "__main__":
    config = VisionConfig()
    analyzer = SDGVisionAnalyzer(config)
    analyzer.run()